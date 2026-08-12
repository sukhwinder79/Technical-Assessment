"""Microsoft Excel tools.

`excel_import_csv` genuinely launches the Excel application through COM
automation and performs a real *Data → From Text* import (a `QueryTable`), then
formats and saves the workbook — this is not a library writing an .xlsx behind
the scenes.

When COM is unavailable (non-Windows host, Docker container, Excel not
installed) the tool degrades gracefully to an `openpyxl` writer, reports
`engine="openpyxl"` and `excel_launched=False`, and tells the model exactly what
happened so it can report the step honestly instead of claiming success.
"""

from __future__ import annotations

import platform
import sys
from contextlib import contextmanager
from pathlib import Path
from typing import Any, Iterator

from pydantic import BaseModel, Field

from ..errors import ToolError
from ..logging_setup import get_logger
from .base import Tool, ToolContext
from .data_tools import ConvertSpreadsheetTool, read_csv

log = get_logger(__name__)

# ---- Excel interop constants (from the Excel object model) -----------------
XL_DELIMITED = 1
XL_TEXT_QUALIFIER_DOUBLE_QUOTE = 1
XL_OVERWRITE_CELLS = 0
XL_OPEN_XML_WORKBOOK = 51          # .xlsx
XL_OPEN_XML_WORKBOOK_MACRO = 52    # .xlsm
CODEPAGE_UTF8 = 65001

# COM HRESULTs worth treating as transient rather than fatal.
_RETRYABLE_HRESULTS = {
    -2147418111,  # RPC_E_CALL_REJECTED  — Excel busy / modal dialog open
    -2146777998,  # VBA_E_IGNORE         — object invoked has disconnected
    -2147417846,  # RPC_E_SERVERCALL_RETRYLATER
}


def _com_available() -> tuple[bool, str]:
    """Can we drive Excel on this host? Returns (available, reason)."""
    if sys.platform != "win32":
        return False, f"Not running on Windows (platform={sys.platform})."
    try:
        import win32com.client  # noqa: F401
    except ImportError:
        return False, "pywin32 is not installed (pip install pywin32)."
    return True, "ok"


@contextmanager
def _com_apartment() -> Iterator[None]:
    """Initialise COM for the calling thread and always tear it down.

    Required because the FastAPI server and the MCP server run tools on worker
    threads, which start with no COM apartment.

    The `gc.collect()` before `CoUninitialize` is load-bearing: any COM pointer
    still reachable when the apartment closes gets its `Release()` deferred to
    an apartment that no longer exists, which surfaces as a process-level
    RPC_E_DISCONNECTED (0x80010108) fault rather than a Python exception.
    """
    import gc

    import pythoncom

    pythoncom.CoInitialize()
    try:
        yield
    finally:
        gc.collect()
        pythoncom.CoUninitialize()


def _wrap_com_error(exc: BaseException, action: str) -> ToolError:
    hresult = getattr(exc, "hresult", None) or getattr(exc, "args", [None])[0]
    retryable = hresult in _RETRYABLE_HRESULTS
    detail = getattr(exc, "excepinfo", None)
    message = f"Excel COM automation failed while {action}: {exc}"
    if retryable:
        remediation = "Excel appears busy. Close any open dialog box in Excel; the agent will retry."
    else:
        remediation = (
            "Verify Microsoft Excel is installed and can be launched manually. "
            "If Excel is unavailable, use convert_spreadsheet with target_format='xlsx' "
            "to produce the workbook and report that Excel could not be launched."
        )
    return ToolError(
        message,
        retryable=retryable,
        remediation=remediation,
        details={"hresult": hresult, "com_detail": str(detail) if detail else None},
    )


# ==========================================================================
#  Tool: excel_probe
# ==========================================================================


def _probe_excel_registration() -> tuple[bool, str, str | None]:
    """Is `Excel.Application` registered for COM automation?

    Reads the registry instead of launching Excel: instant, no side effects, no
    COM apartment needed, and it avoids the shutdown race a launch-then-Quit
    probe hits — Excel begins tearing down on `Quit()`, so releasing the
    interface pointer afterwards can surface RPC_E_DISCONNECTED at the process
    level. Resolving `Excel.Application` -> `CurVer` -> `CLSID` is exactly what
    `CoCreateInstance` does when the agent later dispatches Excel for real.
    """
    import winreg

    try:
        with winreg.OpenKey(winreg.HKEY_CLASSES_ROOT, r"Excel.Application\CurVer") as key:
            prog_id, _ = winreg.QueryValueEx(key, "")
    except OSError as exc:
        return False, f"'Excel.Application' is not registered for COM automation: {exc}", None

    version = prog_id.rsplit(".", 1)[-1] if "." in prog_id else None  # "Excel.Application.16" -> "16"

    try:
        with winreg.OpenKey(winreg.HKEY_CLASSES_ROOT, rf"{prog_id}\CLSID") as key:
            winreg.QueryValueEx(key, "")
    except OSError as exc:
        return False, f"'{prog_id}' has no CLSID registered — Excel may need a repair install: {exc}", version

    detail = f" (Excel {version}.0)" if version else ""
    return True, f"Microsoft Excel is registered for COM automation{detail}.", version


def _probe_excel_application() -> tuple[bool, str, str | None]:
    """Actually start Excel, read its version, and quit. Needs an apartment.

    Stronger evidence than the registry check, at the cost of a few seconds and
    an Excel process. Every COM reference is confined to this frame and dropped
    before it returns, so nothing survives into `CoUninitialize`.
    """
    import win32com.client as win32

    excel = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        version = str(excel.Version)
        return True, f"Microsoft Excel launched successfully (version {version}).", version
    except Exception as exc:  # noqa: BLE001
        return False, f"Excel could not be started: {exc}", None
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:  # pragma: no cover
                pass
            del excel


class ExcelProbeArgs(BaseModel):
    deep_check: bool = Field(
        default=False,
        description=(
            "Actually launch and close Excel instead of only checking that it is registered. "
            "Slower (a few seconds) but proves automation works end to end."
        ),
    )


class ExcelProbeTool(Tool):
    name = "excel_probe"
    description = (
        "Check whether Microsoft Excel can be automated on this machine, and report its version. "
        "By default this is an instant registry check with no side effects; pass deep_check=true to "
        "actually launch and close Excel. Call this before excel_import_csv if you are unsure the "
        "environment supports Excel, so you can plan a fallback. Never guess — this tool answers."
    )
    args_model = ExcelProbeArgs
    tags = ("excel", "probe")
    max_retries = 0

    def run(self, args: ExcelProbeArgs, ctx: ToolContext) -> dict[str, Any]:
        available, reason = _com_available()
        result: dict[str, Any] = {
            "ok": True,
            "platform": platform.platform(),
            "com_available": available,
            "excel_available": False,
            "reason": reason,
        }
        if not available:
            result["fallback"] = "Use convert_spreadsheet(target_format='xlsx') to produce the workbook."
            return result

        if args.deep_check:
            with _com_apartment():
                ok, reason, version = _probe_excel_application()
            result["check"] = "launched Excel"
        else:
            ok, reason, version = _probe_excel_registration()
            result["check"] = "COM registration"

        result["excel_available"] = ok
        result["reason"] = reason
        if version:
            result["excel_version"] = version
        if not ok:
            result["fallback"] = "Use convert_spreadsheet(target_format='xlsx') instead."

        ctx.memory.remember("excel_available", result["excel_available"])
        log.info("excel.probe", **{k: v for k, v in result.items() if k != "ok"})
        return result


# ==========================================================================
#  Tool: excel_import_csv
# ==========================================================================


class ExcelImportCsvArgs(BaseModel):
    csv_path: str = Field(description="Path to the CSV file to import.")
    workbook_filename: str = Field(
        default="employees.xlsx",
        description="Name of the workbook to save (relative paths land in the agent workspace).",
    )
    sheet_name: str = Field(default="Employees", description="Worksheet name for the imported data.")
    apply_formatting: bool = Field(
        default=True,
        description="Bold header row, freeze panes, auto-fit columns, add an AutoFilter and format numbers.",
    )

    # NOTE: window visibility and whether Excel stays open afterwards are
    # deliberately NOT exposed to the model. They are properties of the host and
    # of the human watching the screen, and the agent has no basis for judging
    # either — when they were arguments, the model kept choosing
    # `visible=false, keep_open=false`, silently overriding the operator's
    # EXCEL_VISIBLE/EXCEL_KEEP_OPEN settings and hiding the very thing a demo is
    # meant to show. Leaving them out also removes two nullable fields from the
    # schema sent on every turn.


class ExcelImportCsvTool(Tool):
    name = "excel_import_csv"
    description = (
        "Launch the Microsoft Excel application, import a CSV file into a new worksheet using Excel's "
        "own text-import engine, apply formatting, and save the result as an .xlsx workbook. "
        "This is the tool to use when the user asks to 'open Excel' or 'import into Excel'. "
        "If Excel cannot be launched the tool falls back to writing the .xlsx directly and reports "
        "excel_launched=false — report that honestly rather than claiming Excel was opened."
    )
    args_model = ExcelImportCsvArgs
    tags = ("excel", "import", "write")
    max_retries = 2

    def run(self, args: ExcelImportCsvArgs, ctx: ToolContext) -> dict[str, Any]:
        csv_path = ctx.resolve(args.csv_path)
        headers, rows = read_csv(csv_path)  # validates existence + non-empty

        workbook_path = ctx.resolve(args.workbook_filename)
        if workbook_path.suffix.lower() not in (".xlsx", ".xlsm"):
            workbook_path = workbook_path.with_suffix(".xlsx")
        workbook_path.parent.mkdir(parents=True, exist_ok=True)

        # Operator configuration, not an agent decision — see ExcelImportCsvArgs.
        visible = ctx.settings.excel_visible
        keep_open = ctx.settings.excel_keep_open

        available, reason = _com_available()
        if not available:
            return self._fallback(csv_path, workbook_path, args, ctx, headers, rows, reason)

        ctx.events.emit(
            "tool_started",
            f"Launching Microsoft Excel and importing {csv_path.name}…",
            tool=self.name,
            phase="excel_launch",
        )

        with _com_apartment():
            result = self._import_via_com(
                csv_path=csv_path,
                workbook_path=workbook_path,
                sheet_name=args.sheet_name,
                visible=visible,
                keep_open=keep_open,
                apply_formatting=args.apply_formatting,
                headers=headers,
                row_count=len(rows),
            )

        ctx.memory.remember("last_workbook_path", str(workbook_path))
        ctx.memory.remember("excel_available", True)
        log.info("excel.import.done", **{k: v for k, v in result.items() if k != "ok"})
        return result

    # ---- COM path ----------------------------------------------------------

    def _import_via_com(
        self,
        *,
        csv_path: Path,
        workbook_path: Path,
        sheet_name: str,
        visible: bool,
        keep_open: bool,
        apply_formatting: bool,
        headers: list[str],
        row_count: int,
    ) -> dict[str, Any]:
        import win32com.client as win32

        excel = None
        workbook = None
        sheet = None
        query_table = None
        used = None
        try:
            # DispatchEx spins up a dedicated Excel instance so we never disturb
            # a workbook the user already has open.
            excel = win32.DispatchEx("Excel.Application")
        except Exception as exc:  # noqa: BLE001
            raise _wrap_com_error(exc, "starting Excel") from exc

        try:
            excel.Visible = bool(visible)
            excel.DisplayAlerts = False
            excel.ScreenUpdating = True

            workbook = excel.Workbooks.Add()
            sheet = workbook.Worksheets(1)
            sheet.Name = (sheet_name or "Employees")[:31]

            # --- The actual import: Excel's own delimited-text engine --------
            query_table = sheet.QueryTables.Add(
                Connection=f"TEXT;{csv_path}",
                Destination=sheet.Range("A1"),
            )
            query_table.TextFilePlatform = CODEPAGE_UTF8
            query_table.TextFileParseType = XL_DELIMITED
            query_table.TextFileTextQualifier = XL_TEXT_QUALIFIER_DOUBLE_QUOTE
            query_table.TextFileConsecutiveDelimiter = False
            query_table.TextFileTabDelimiter = False
            query_table.TextFileSemicolonDelimiter = False
            query_table.TextFileCommaDelimiter = True
            query_table.TextFileSpaceDelimiter = False
            query_table.RefreshStyle = XL_OVERWRITE_CELLS
            query_table.AdjustColumnWidth = True
            query_table.PreserveFormatting = True
            query_table.Refresh(BackgroundQuery=False)

            # Drop the external-data link so the saved file is plain data.
            try:
                query_table.Delete()
            except Exception:  # pragma: no cover - Excel version differences
                pass
            query_table = None
            self._drop_connections(workbook)

            used = sheet.UsedRange
            imported_rows = int(used.Rows.Count)
            imported_cols = int(used.Columns.Count)
            used = None

            if apply_formatting:
                self._format_sheet(excel, sheet, headers, imported_rows, imported_cols)

            if workbook_path.exists():
                try:
                    workbook_path.unlink()
                except OSError as exc:
                    raise ToolError(
                        f"Cannot overwrite {workbook_path}: {exc}",
                        remediation="The workbook may be open in another Excel window. Close it or pick a new name.",
                    ) from exc

            file_format = (
                XL_OPEN_XML_WORKBOOK_MACRO
                if workbook_path.suffix.lower() == ".xlsm"
                else XL_OPEN_XML_WORKBOOK
            )
            workbook.SaveAs(str(workbook_path), FileFormat=file_format)

            excel_version = str(excel.Version)

            if not keep_open:
                # Release strictly inside-out — worksheet, then workbook, then
                # the application. Any child pointer still held when Excel exits
                # gets its Release() delivered to a dead server, which surfaces
                # as an RPC_E_DISCONNECTED process fault rather than an exception.
                sheet = None
                workbook.Close(SaveChanges=False)
                workbook = None
                excel.Quit()
                excel = None

            return {
                "ok": True,
                "engine": "excel-com",
                "excel_launched": True,
                "excel_version": excel_version,
                "excel_left_open": keep_open,
                "excel_visible": bool(visible),
                "csv_path": str(csv_path),
                "workbook_path": str(workbook_path),
                "sheet_name": sheet_name[:31],
                "rows_written": imported_rows,          # includes the header row
                "data_rows": max(imported_rows - 1, 0),
                "columns_written": imported_cols,
                "expected_data_rows": row_count,
                "size_bytes": workbook_path.stat().st_size,
                "formatting_applied": apply_formatting,
            }

        except ToolError:
            raise
        except Exception as exc:  # noqa: BLE001
            raise _wrap_com_error(exc, "importing the CSV") from exc
        finally:
            # On any failure make sure we never leave a headless Excel process behind.
            if excel is not None and not keep_open:
                try:
                    used = None
                    query_table = None
                    sheet = None
                    if workbook is not None:
                        workbook.Close(SaveChanges=False)
                        workbook = None
                    excel.Quit()
                    excel = None
                except Exception:  # pragma: no cover
                    pass

    @staticmethod
    def _drop_connections(workbook) -> None:
        try:
            for index in range(workbook.Connections.Count, 0, -1):
                workbook.Connections.Item(index).Delete()
        except Exception:  # pragma: no cover - not all Excel builds expose this
            pass

    @staticmethod
    def _format_sheet(excel, sheet, headers: list[str], rows: int, cols: int) -> None:
        header = sheet.Range(sheet.Cells(1, 1), sheet.Cells(1, cols))
        header.Font.Bold = True
        header.Font.Color = 0xFFFFFF          # white (BGR)
        header.Interior.Color = 0x97522F      # dark blue in BGR
        header.HorizontalAlignment = -4108    # xlCenter

        # Number formats driven by the header names, not fixed positions.
        for index, name in enumerate(headers[:cols], start=1):
            lowered = str(name).lower()
            if "salary" in lowered or "compensation" in lowered:
                fmt = "#,##0"
            elif "date" in lowered:
                fmt = "yyyy-mm-dd"
            elif "rating" in lowered:
                fmt = "0.0"
            else:
                continue
            sheet.Range(sheet.Cells(2, index), sheet.Cells(max(rows, 2), index)).NumberFormat = fmt

        sheet.Columns.AutoFit()
        sheet.Range("A2").Select()
        excel.ActiveWindow.FreezePanes = True
        try:
            if not sheet.AutoFilterMode:
                sheet.Range(sheet.Cells(1, 1), sheet.Cells(rows, cols)).AutoFilter(1)
        except Exception:  # pragma: no cover
            pass
        sheet.Range("A1").Select()

    # ---- fallback path -----------------------------------------------------

    def _fallback(
        self,
        csv_path: Path,
        workbook_path: Path,
        args: ExcelImportCsvArgs,
        ctx: ToolContext,
        headers: list[str],
        rows: list[list[str]],
        reason: str,
    ) -> dict[str, Any]:
        ctx.events.emit(
            "tool_started",
            f"Excel is unavailable ({reason}) — writing the workbook with openpyxl instead.",
            tool=self.name,
            phase="excel_fallback",
        )
        ConvertSpreadsheetTool._write_xlsx(workbook_path, headers, rows, args.sheet_name)
        ctx.memory.remember("last_workbook_path", str(workbook_path))
        ctx.memory.remember("excel_available", False)
        log.warning("excel.fallback", reason=reason, workbook=str(workbook_path))
        return {
            "ok": True,
            "engine": "openpyxl",
            "excel_launched": False,
            "excel_unavailable_reason": reason,
            "csv_path": str(csv_path),
            "workbook_path": str(workbook_path),
            "sheet_name": args.sheet_name[:31],
            "rows_written": len(rows) + 1,
            "data_rows": len(rows),
            "columns_written": len(headers),
            "size_bytes": workbook_path.stat().st_size,
            "note": (
                "The workbook was created successfully, but the Microsoft Excel application was "
                "NOT launched on this host. Report this step as completed with a fallback engine."
            ),
        }


# ==========================================================================
#  Tool: excel_verify_workbook
# ==========================================================================


class ExcelVerifyWorkbookArgs(BaseModel):
    workbook_path: str = Field(description="Path to the .xlsx workbook to verify.")
    expected_row_count: int | None = Field(
        default=None,
        description="Expected number of DATA rows (excluding the header). Omit to skip the count check.",
    )
    expected_columns: list[str] | None = Field(
        default=None, description="Expected header names, in order. Omit to skip the header check."
    )
    sheet_name: str | None = Field(default=None, description="Worksheet to verify. Defaults to the first sheet.")


class ExcelVerifyWorkbookTool(Tool):
    name = "excel_verify_workbook"
    description = (
        "Re-open a saved .xlsx workbook from disk and confirm the data actually landed: sheet names, "
        "header row, row count and a sample of the data. Call this after excel_import_csv so you can "
        "report verified success rather than assumed success."
    )
    args_model = ExcelVerifyWorkbookArgs
    tags = ("excel", "verify")
    max_retries = 1

    def run(self, args: ExcelVerifyWorkbookArgs, ctx: ToolContext) -> dict[str, Any]:
        from openpyxl import load_workbook

        path = ctx.resolve(args.workbook_path)
        if not path.exists():
            raise ToolError(
                f"Workbook not found: {path}",
                remediation="Run excel_import_csv first, or check the path.",
            )

        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise ToolError(
                f"Could not open workbook {path}: {exc}",
                retryable=True,
                remediation="The file may still be being written by Excel. Retrying may help.",
            ) from exc

        try:
            sheet_names = list(workbook.sheetnames)
            sheet = workbook[args.sheet_name] if args.sheet_name in sheet_names else workbook[sheet_names[0]]
            values = [list(row) for row in sheet.iter_rows(values_only=True)]
        finally:
            workbook.close()

        values = [row for row in values if any(cell is not None and str(cell) != "" for cell in row)]
        if not values:
            raise ToolError(f"Workbook {path} contains no data.")

        headers = [str(cell) if cell is not None else "" for cell in values[0]]
        data_rows = len(values) - 1

        checks: dict[str, Any] = {}
        problems: list[str] = []

        if args.expected_row_count is not None:
            passed = data_rows == args.expected_row_count
            checks["row_count"] = {"expected": args.expected_row_count, "actual": data_rows, "passed": passed}
            if not passed:
                problems.append(f"Expected {args.expected_row_count} data rows but found {data_rows}.")

        if args.expected_columns is not None:
            passed = [h.strip() for h in headers] == [c.strip() for c in args.expected_columns]
            checks["columns"] = {"expected": args.expected_columns, "actual": headers, "passed": passed}
            if not passed:
                problems.append("Header row does not match the expected columns.")

        verified = not problems
        log.info("excel.verify", path=str(path), rows=data_rows, verified=verified)

        return {
            "ok": True,
            "verified": verified,
            "workbook_path": str(path),
            "sheet_names": sheet_names,
            "sheet_verified": sheet.title,
            "columns": headers,
            "data_row_count": data_rows,
            "checks": checks,
            "problems": problems,
            "sample_rows": [[_render_cell(cell) for cell in row] for row in values[1:4]],
        }


def _render_cell(value: Any) -> str:
    """Stringify a cell for the sample preview.

    Excel's text import turns ISO date strings into real datetimes (which is
    what we want — they sort and filter correctly), but `str(datetime)` would
    show a pointless midnight component in the report.
    """
    import datetime as dt

    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat() if value.time() == dt.time.min else value.isoformat(sep=" ")
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value)
