"""Data tools: generate sample employee CSVs, convert between spreadsheet
formats (CSV / XLSX / ODS), and preview a file before importing it."""

from __future__ import annotations

import csv
import datetime as dt
import random
import unicodedata
from pathlib import Path
from typing import Any, Literal

from pydantic import BaseModel, Field

from ..errors import ToolError
from ..logging_setup import get_logger
from .base import Tool, ToolContext

log = get_logger(__name__)

# --------------------------------------------------------------------------
# Sample data vocabulary — realistic rather than lorem-ipsum, because the whole
# point of the deliverable is that a human opens the sheet and it looks real.
# --------------------------------------------------------------------------

FIRST_NAMES = [
    "John", "Alice", "Priya", "Marcus", "Sofia", "Daniel", "Aisha", "Liam",
    "Mei", "Carlos", "Hannah", "Omar", "Elena", "Noah", "Fatima", "Ethan",
    "Ananya", "Lucas", "Grace", "Yusuf", "Isabella", "Rohan", "Chloe", "Andre",
    "Nadia", "Thomas", "Leila", "Victor", "Amara", "Kenji",
]

LAST_NAMES = [
    "Smith", "Brown", "Sharma", "Reyes", "Novak", "OConnor", "Khan", "Patel",
    "Chen", "Mendoza", "Lindqvist", "Haddad", "Petrova", "Walker", "Rahman",
    "Fitzgerald", "Iyer", "Moreau", "Bennett", "Osei", "Rossi", "Kapoor",
    "Dubois", "Silva", "Ahmed", "Muller", "Farah", "Kowalski", "Diallo", "Tanaka",
]

DEPARTMENTS: dict[str, tuple[list[str], tuple[int, int]]] = {
    "Engineering": (
        ["Software Engineer", "Senior Software Engineer", "QA Engineer", "DevOps Engineer", "Engineering Manager"],
        (78_000, 165_000),
    ),
    "Sales": (
        ["Account Executive", "Sales Development Rep", "Regional Sales Manager", "Solutions Consultant"],
        (52_000, 130_000),
    ),
    "Marketing": (
        ["Marketing Specialist", "Content Strategist", "Growth Marketer", "Marketing Manager"],
        (55_000, 118_000),
    ),
    "HR": (
        ["HR Generalist", "Recruiter", "People Operations Lead", "HR Business Partner"],
        (54_000, 112_000),
    ),
    "Finance": (
        ["Financial Analyst", "Accountant", "Payroll Specialist", "Finance Manager"],
        (58_000, 135_000),
    ),
    "Operations": (
        ["Operations Analyst", "Logistics Coordinator", "Operations Manager", "Facilities Lead"],
        (48_000, 110_000),
    ),
    "Customer Support": (
        ["Support Specialist", "Technical Support Engineer", "Support Team Lead"],
        (42_000, 92_000),
    ),
}

LOCATIONS = [
    "New York, NY", "Austin, TX", "San Francisco, CA", "Chicago, IL",
    "Remote (US)", "London, UK", "Bengaluru, IN", "Berlin, DE",
    "Toronto, CA", "Singapore, SG",
]

EMPLOYMENT_TYPES = ["Full-time", "Full-time", "Full-time", "Contract", "Part-time"]

BASE_COLUMNS = ["Employee ID", "Name", "Department", "Email", "Salary"]
EXTENDED_COLUMNS = ["Job Title", "Location", "Hire Date", "Employment Type", "Performance Rating"]


def _slug(text: str) -> str:
    """ASCII-safe local part for an email address."""
    normalised = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    return "".join(ch for ch in normalised.lower() if ch.isalnum())


def generate_employee_rows(
    row_count: int,
    *,
    seed: int | None = None,
    include_extended: bool = True,
    email_domain: str = "example.com",
) -> tuple[list[str], list[list[Any]]]:
    """Build the header + data rows for a sample employee dataset.

    Pure function: no I/O, deterministic when `seed` is provided. That makes it
    straightforward to assert on in tests.
    """
    rng = random.Random(seed)
    headers = BASE_COLUMNS + (EXTENDED_COLUMNS if include_extended else [])

    today = dt.date.today()
    used_emails: set[str] = set()
    rows: list[list[Any]] = []

    for index in range(1, row_count + 1):
        first, last = rng.choice(FIRST_NAMES), rng.choice(LAST_NAMES)
        name = f"{first} {last}"
        department = rng.choice(list(DEPARTMENTS))
        titles, (low, high) = DEPARTMENTS[department]

        local = f"{_slug(first)}.{_slug(last)}"
        if local in used_emails:
            local = f"{local}{index}"
        used_emails.add(local)

        salary = rng.randrange(low, high + 1, 500)
        row: list[Any] = [
            f"EMP{index:03d}",
            name,
            department,
            f"{local}@{email_domain}",
            salary,
        ]

        if include_extended:
            hire_date = today - dt.timedelta(days=rng.randint(45, 8 * 365))
            row += [
                rng.choice(titles),
                rng.choice(LOCATIONS),
                hire_date.isoformat(),
                rng.choice(EMPLOYMENT_TYPES),
                round(rng.uniform(2.8, 5.0), 1),
            ]

        rows.append(row)

    return headers, rows


def write_csv(path: Path, headers: list[str], rows: list[list[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    # utf-8-sig so Excel on Windows detects UTF-8 without a manual import step.
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


#: A trailing empty ODS cell can claim to repeat tens of thousands of times.
#: Expanding that verbatim would build a useless megabyte-wide row.
MAX_ODS_REPEAT = 256


def read_ods(path: Path) -> tuple[list[str], list[list[str]]]:
    """Read the first table of an OpenDocument spreadsheet.

    Needed so ODS is a first-class format rather than write-only: it lets
    `convert_spreadsheet` take an .ods as input, and lets the API preview one
    (which also doubles as a check that the ODS we wrote is well formed).
    """
    try:
        from odf.opendocument import load
        from odf.table import Table, TableCell, TableRow
        from odf.text import P
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise ToolError(
            "ODS support requires the 'odfpy' package.", remediation="pip install odfpy"
        ) from exc

    try:
        document = load(str(path))
    except Exception as exc:  # noqa: BLE001
        raise ToolError(f"Could not open ODS file {path}: {exc}") from exc

    tables = document.spreadsheet.getElementsByType(Table)
    if not tables:
        raise ToolError(f"ODS file has no tables: {path}")

    parsed: list[list[str]] = []
    for table_row in tables[0].getElementsByType(TableRow):
        values: list[str] = []
        for cell in table_row.getElementsByType(TableCell):
            repeat = int(cell.getAttribute("numbercolumnsrepeated") or 1)
            text = "".join(str(node) for node in cell.getElementsByType(P))
            values.extend([text] * min(repeat, MAX_ODS_REPEAT))
        while values and values[-1] == "":       # drop the trailing padding
            values.pop()
        if values:
            parsed.append(values)

    if not parsed:
        raise ToolError(f"ODS file contains no data: {path}")
    return parsed[0], parsed[1:]


def read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    if not path.exists():
        raise ToolError(
            f"CSV not found: {path}",
            remediation="Call generate_employee_csv first, or pass the correct csv_path.",
        )
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = list(csv.reader(handle))
    if not reader:
        raise ToolError(f"CSV is empty: {path}")
    return reader[0], reader[1:]


# ==========================================================================
#  Tool: generate_employee_csv
# ==========================================================================


class GenerateEmployeeCsvArgs(BaseModel):
    row_count: int = Field(
        default=25,
        ge=1,
        le=5000,
        description="Number of employee records to generate. The assessment asks for at least 20.",
    )
    filename: str = Field(
        default="employees.csv",
        description="Output file name (relative paths land in the agent workspace).",
    )
    include_extended_columns: bool = Field(
        default=True,
        description=(
            "Include Job Title, Location, Hire Date, Employment Type and Performance Rating "
            "in addition to the five required columns."
        ),
    )
    email_domain: str = Field(default="example.com", description="Domain used for generated emails.")
    seed: int | None = Field(
        default=None,
        description="Random seed. Pass a value when the user asks for reproducible data.",
    )


class GenerateEmployeeCsvTool(Tool):
    name = "generate_employee_csv"
    description = (
        "Generate a CSV file of realistic sample employee records and write it to disk. "
        "Columns always include Employee ID, Name, Department, Email and Salary. "
        "Use this whenever the user asks for sample, test, demo or employee data and no "
        "source file was supplied. Returns the absolute path, row count and a preview."
    )
    args_model = GenerateEmployeeCsvArgs
    tags = ("data", "generate")

    def run(self, args: GenerateEmployeeCsvArgs, ctx: ToolContext) -> dict[str, Any]:
        path = ctx.resolve(args.filename)
        if path.suffix.lower() != ".csv":
            path = path.with_suffix(".csv")

        headers, rows = generate_employee_rows(
            args.row_count,
            seed=args.seed,
            include_extended=args.include_extended_columns,
            email_domain=args.email_domain,
        )

        try:
            write_csv(path, headers, rows)
        except OSError as exc:
            raise ToolError(
                f"Could not write CSV to {path}: {exc}",
                remediation="Choose a different filename or check directory permissions.",
            ) from exc

        ctx.memory.remember("last_csv_path", str(path))
        ctx.memory.remember("last_csv_row_count", len(rows))
        ctx.memory.remember("last_csv_columns", headers)

        log.info("csv.generated", path=str(path), rows=len(rows), columns=len(headers))

        return {
            "ok": True,
            "csv_path": str(path),
            "row_count": len(rows),
            "column_count": len(headers),
            "columns": headers,
            "size_bytes": path.stat().st_size,
            "preview": [headers, *[[str(cell) for cell in row] for row in rows[:3]]],
        }


# ==========================================================================
#  Tool: convert_spreadsheet
# ==========================================================================


class ConvertSpreadsheetArgs(BaseModel):
    source_path: str = Field(description="Path to the source CSV or XLSX file.")
    target_format: Literal["xlsx", "csv", "ods"] = Field(
        description="Output format. xlsx = Excel workbook, ods = OpenDocument spreadsheet."
    )
    output_filename: str | None = Field(
        default=None,
        description="Optional output file name. Defaults to the source name with the new extension.",
    )
    sheet_name: str = Field(default="Employees", description="Worksheet name for xlsx/ods output.")


class ConvertSpreadsheetTool(Tool):
    name = "convert_spreadsheet"
    description = (
        "Convert a spreadsheet between CSV, XLSX and ODS without needing Microsoft Excel. "
        "Reads .csv, .xlsx and .ods; writes any of the three. "
        "Use this when the user asks for an extra format (for example 'also give me an .ods'), "
        "or as a fallback to produce an .xlsx when Excel automation is unavailable. "
        "Note: this writes a file directly and does NOT open the Excel application."
    )
    args_model = ConvertSpreadsheetArgs
    tags = ("data", "convert")

    def run(self, args: ConvertSpreadsheetArgs, ctx: ToolContext) -> dict[str, Any]:
        source = ctx.resolve(args.source_path)
        if not source.exists():
            raise ToolError(
                f"Source file not found: {source}",
                remediation="Generate or locate the source file first.",
            )

        if args.output_filename:
            target = ctx.resolve(args.output_filename)
        else:
            target = source.with_suffix(f".{args.target_format}")
        target = target.with_suffix(f".{args.target_format}")
        target.parent.mkdir(parents=True, exist_ok=True)

        headers, rows = self._load(source)

        if args.target_format == "csv":
            write_csv(target, headers, rows)
        elif args.target_format == "xlsx":
            self._write_xlsx(target, headers, rows, args.sheet_name)
        else:
            self._write_ods(target, headers, rows, args.sheet_name)

        ctx.memory.remember(f"last_{args.target_format}_path", str(target))
        log.info("spreadsheet.converted", source=str(source), target=str(target), rows=len(rows))

        return {
            "ok": True,
            "source_path": str(source),
            "output_path": str(target),
            "format": args.target_format,
            "row_count": len(rows),
            "column_count": len(headers),
            "size_bytes": target.stat().st_size,
        }

    # ---- format helpers ----------------------------------------------------

    @staticmethod
    def _load(source: Path) -> tuple[list[str], list[list[Any]]]:
        suffix = source.suffix.lower()
        if suffix == ".csv":
            return read_csv(source)
        if suffix in (".xlsx", ".xlsm"):
            from openpyxl import load_workbook

            workbook = load_workbook(source, read_only=True, data_only=True)
            sheet = workbook.active
            values = [list(row) for row in sheet.iter_rows(values_only=True)]
            workbook.close()
            if not values:
                raise ToolError(f"Workbook has no data: {source}")
            headers = [str(cell) if cell is not None else "" for cell in values[0]]
            return headers, [list(row) for row in values[1:]]
        if suffix == ".ods":
            return read_ods(source)
        raise ToolError(
            f"Unsupported source format '{suffix}'.",
            remediation="Supported inputs are .csv, .xlsx and .ods.",
        )

    @staticmethod
    def _write_xlsx(target: Path, headers: list[str], rows: list[list[Any]], sheet_name: str) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name[:31] or "Sheet1"
        sheet.append(headers)
        for row in rows:
            sheet.append([_coerce_number(cell) for cell in row])

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", start_color="2F5597")
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        _apply_number_formats(sheet, headers)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        for index, header in enumerate(headers, start=1):
            longest = max(
                [len(str(header))] + [len(str(row[index - 1])) for row in rows if index - 1 < len(row)] or [10]
            )
            sheet.column_dimensions[get_column_letter(index)].width = min(max(longest + 3, 12), 40)

        workbook.save(target)

    @staticmethod
    def _write_ods(target: Path, headers: list[str], rows: list[list[Any]], sheet_name: str) -> None:
        try:
            from odf.opendocument import OpenDocumentSpreadsheet
            from odf.table import Table, TableCell, TableRow
            from odf.text import P
        except ImportError as exc:  # pragma: no cover - dependency guard
            raise ToolError(
                "ODS support requires the 'odfpy' package.",
                remediation="pip install odfpy",
            ) from exc

        document = OpenDocumentSpreadsheet()
        table = Table(name=sheet_name[:31] or "Sheet1")

        def add_row(values: list[Any]) -> None:
            table_row = TableRow()
            for value in values:
                number = _coerce_number(value)
                if isinstance(number, (int, float)) and not isinstance(number, bool):
                    cell = TableCell(valuetype="float", value=float(number))
                else:
                    cell = TableCell(valuetype="string")
                cell.addElement(P(text="" if value is None else str(value)))
                table_row.addElement(cell)
            table.addElement(table_row)

        add_row(headers)
        for row in rows:
            add_row(row)

        document.spreadsheet.addElement(table)
        document.save(str(target))


def _coerce_number(value: Any) -> Any:
    """CSV gives us strings; write real numbers so Excel/Sheets can sum them."""
    if value is None or isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if not text:
        return text
    try:
        return int(text)
    except ValueError:
        pass
    try:
        return float(text)
    except ValueError:
        return value


def _apply_number_formats(sheet, headers: list[str]) -> None:
    from openpyxl.utils import get_column_letter

    for index, header in enumerate(headers, start=1):
        lowered = header.lower()
        if "salary" in lowered or "compensation" in lowered:
            fmt = '#,##0'
        elif "date" in lowered:
            fmt = "yyyy-mm-dd"
        elif "rating" in lowered:
            fmt = "0.0"
        else:
            continue
        letter = get_column_letter(index)
        for cell in sheet[letter][1:]:
            cell.number_format = fmt


# ==========================================================================
#  Tool: read_csv_preview
# ==========================================================================


class ReadCsvPreviewArgs(BaseModel):
    csv_path: str = Field(description="Path to the CSV file to inspect.")
    max_rows: int = Field(default=5, ge=1, le=50, description="How many data rows to return.")


class ReadCsvPreviewTool(Tool):
    name = "read_csv_preview"
    description = (
        "Read the header and the first few rows of a CSV file, plus its total row count. "
        "Use this to inspect a file the user supplied, or to sanity-check data before importing."
    )
    args_model = ReadCsvPreviewArgs
    tags = ("data", "inspect")

    def run(self, args: ReadCsvPreviewArgs, ctx: ToolContext) -> dict[str, Any]:
        path = ctx.resolve(args.csv_path)
        headers, rows = read_csv(path)
        return {
            "ok": True,
            "csv_path": str(path),
            "columns": headers,
            "row_count": len(rows),
            "preview": [headers, *rows[: args.max_rows]],
        }
