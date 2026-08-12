"""FastAPI surface.

    GET  /                 the web UI (single self-contained page)
    POST /runs             start a run, return immediately (202)
    GET  /runs/{id}        poll status / fetch the full report
    GET  /runs/{id}/events Server-Sent Events — the same progress stream the CLI renders
    POST /runs/sync        run and block until finished (handy for curl demos)
    GET  /tools            the live toolbox
    GET  /health

The agent itself is synchronous, so each run executes on a worker thread and
publishes events back onto the event loop with `call_soon_threadsafe`.

The UI is the fourth subscriber to the same event bus the CLI, the SSE stream
and the structured logs already read — it renders the run live rather than
polling for a result.
"""

from __future__ import annotations

import asyncio
import uuid
from pathlib import Path
from typing import Any, Optional

from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from sse_starlette.sse import EventSourceResponse

from .. import __version__
from ..agent import build_agent
from ..config import get_settings
from ..errors import ConfigurationError
from ..events import Event, EventBus
from ..logging_setup import get_logger
from ..tools.registry import build_default_registry

log = get_logger(__name__)

app = FastAPI(
    title="Agentic Spreadsheet Agent",
    version=__version__,
    description=(
        "An autonomous agent that generates employee data, imports it into Microsoft Excel "
        "via COM automation, and uploads it to Google Sheets — from one natural-language instruction."
    ),
)

STATIC_DIR = Path(__file__).parent / "static"
if STATIC_DIR.is_dir():
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


# ---------------------------------------------------------------------------
#  Schemas
# ---------------------------------------------------------------------------


class RunRequest(BaseModel):
    instruction: str = Field(
        default="Create a sample employee CSV and import it into Excel and Google Sheets.",
        description="Natural-language instruction for the agent.",
    )
    session_id: Optional[str] = Field(default=None, description="Reuse a named session for memory.")
    continue_session: bool = Field(default=False, description="Continue the session's conversation.")
    provider: Optional[str] = Field(default=None, description="Override LLM_PROVIDER (e.g. 'groq').")
    model: Optional[str] = Field(default=None, description="Override LLM_MODEL for this run.")
    effort: Optional[str] = Field(default=None, description="Anthropic only: low | medium | high | xhigh | max")
    planning: Optional[bool] = Field(default=None, description="Override the planning pass.")


class RunAccepted(BaseModel):
    run_id: str
    status: str
    events_url: str
    result_url: str


# ---------------------------------------------------------------------------
#  Run bookkeeping
# ---------------------------------------------------------------------------


class RunHandle:
    def __init__(self, run_id: str, request: RunRequest) -> None:
        self.run_id = run_id
        self.request = request
        self.status = "queued"
        self.result: dict[str, Any] | None = None
        self.error: str | None = None
        self.events: list[dict[str, Any]] = []
        self.queues: list[asyncio.Queue] = []
        self.loop: asyncio.AbstractEventLoop | None = None

    # Called from the worker thread.
    def publish(self, event: Event) -> None:
        payload = event.to_dict()
        self.events.append(payload)
        if self.loop is None:
            return
        for queue in list(self.queues):
            self.loop.call_soon_threadsafe(queue.put_nowait, payload)

    def close_streams(self) -> None:
        if self.loop is None:
            return
        for queue in list(self.queues):
            self.loop.call_soon_threadsafe(queue.put_nowait, None)

    def subscribe(self) -> asyncio.Queue:
        queue: asyncio.Queue = asyncio.Queue()
        # Replay what already happened so a late subscriber sees the whole run.
        for payload in self.events:
            queue.put_nowait(payload)
        if self.status in ("completed", "partial", "failed"):
            queue.put_nowait(None)
        self.queues.append(queue)
        return queue


RUNS: dict[str, RunHandle] = {}


def _execute(handle: RunHandle) -> None:
    """Blocking agent execution; runs on a thread-pool worker."""
    request = handle.request
    settings = get_settings()
    if request.provider:
        settings.llm_provider = request.provider  # type: ignore[assignment]
    if request.model:
        settings.llm_model = request.model
    if request.effort:
        settings.agent_effort = request.effort  # type: ignore[assignment]
    if request.planning is not None:
        settings.agent_planning = request.planning

    bus = EventBus()
    bus.subscribe(handle.publish)

    handle.status = "running"
    try:
        agent = build_agent(
            settings=settings,
            events=bus,
            session_id=request.session_id or f"api-{uuid.uuid4().hex[:8]}",
        )
        result = agent.run(
            request.instruction,
            continue_session=request.continue_session,
            run_id=handle.run_id,
        )
        handle.result = result.to_dict()
        handle.status = result.status
    except ConfigurationError as exc:
        handle.status = "failed"
        handle.error = str(exc)
        log.error("api.run.misconfigured", error=str(exc))
    except Exception as exc:  # noqa: BLE001
        handle.status = "failed"
        handle.error = f"{type(exc).__name__}: {exc}"
        log.exception("api.run.crashed")
    finally:
        handle.close_streams()


# ---------------------------------------------------------------------------
#  Endpoints
# ---------------------------------------------------------------------------


@app.get("/", include_in_schema=False, summary="Web UI")
async def index() -> FileResponse:
    page = STATIC_DIR / "index.html"
    if not page.is_file():  # pragma: no cover - only if package data is missing
        raise HTTPException(status_code=404, detail="UI assets are not installed.")
    # No-store so a redeploy is never served from the browser cache mid-demo.
    return FileResponse(page, media_type="text/html", headers={"cache-control": "no-store"})


@app.get("/health", summary="Liveness + environment summary")
async def health() -> dict[str, Any]:
    from ..tools.excel_tools import _com_available

    settings = get_settings()
    excel_available, excel_reason = _com_available()
    return {
        "status": "ok",
        "version": __version__,
        "llm_provider": settings.resolved_provider(),
        "model": settings.resolved_llm_model(),
        "api_key_configured": bool(settings.resolved_llm_api_key()),
        "excel_com_available": excel_available,
        "excel_note": excel_reason,
        "google_auth_mode": settings.google_auth_mode,
        "active_runs": sum(1 for h in RUNS.values() if h.status == "running"),
    }


# ---------------------------------------------------------------------------
#  Generated files
#
#  The agent writes real files to disk, so the UI needs a way to show them.
#  Everything here is confined to WORKSPACE_DIR: a browser must never be able
#  to walk out of it with `..` or an absolute path.
# ---------------------------------------------------------------------------

PREVIEWABLE = {".csv", ".xlsx", ".xlsm", ".ods"}


def _workspace() -> Path:
    return Path(get_settings().workspace_dir).resolve()


def _safe_file(name: str) -> Path:
    """Resolve `name` inside the workspace, or 404. Never escapes it."""
    workspace = _workspace()
    # Take the basename only: a path separator or drive letter cannot survive.
    candidate = (workspace / Path(name).name).resolve()
    if candidate.parent != workspace or not candidate.is_file():
        raise HTTPException(status_code=404, detail=f"No such file in the workspace: {name}")
    return candidate


@app.get("/files", summary="List files the agent has generated")
async def list_files() -> dict[str, Any]:
    workspace = _workspace()
    if not workspace.is_dir():
        return {"workspace": str(workspace), "files": []}
    files = [
        {
            "name": path.name,
            "size_bytes": path.stat().st_size,
            "modified": path.stat().st_mtime,
            "previewable": path.suffix.lower() in PREVIEWABLE,
            "download_url": f"/files/{path.name}",
            "preview_url": f"/files/{path.name}/preview",
        }
        for path in sorted(workspace.iterdir(), key=lambda p: p.stat().st_mtime, reverse=True)
        if path.is_file() and not path.name.startswith("~$")  # skip Excel lock files
    ]
    return {"workspace": str(workspace), "files": files}


@app.get("/files/{name}", summary="Download a generated file")
async def download_file(name: str) -> FileResponse:
    path = _safe_file(name)
    return FileResponse(path, filename=path.name, media_type="application/octet-stream")


@app.get("/files/{name}/preview", summary="Preview a generated spreadsheet as rows")
async def preview_file(name: str, rows: int = 25) -> dict[str, Any]:
    """Read a CSV/XLSX/ODS back and return its contents as JSON.

    This is what lets the browser show the *contents* of the workbook Excel
    produced, rather than just its path.
    """
    path = _safe_file(name)
    suffix = path.suffix.lower()
    rows = max(1, min(rows, 200))

    if suffix == ".csv":
        from ..tools.data_tools import read_csv

        headers, data = read_csv(path)
    elif suffix == ".ods":
        from ..tools.data_tools import read_ods

        headers, data = read_ods(path)
    elif suffix in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            values = [list(r) for r in sheet.iter_rows(values_only=True)]
        finally:
            workbook.close()
        values = [r for r in values if any(c is not None and str(c) != "" for c in r)]
        if not values:
            raise HTTPException(status_code=422, detail="The workbook has no data.")
        headers = [str(c) if c is not None else "" for c in values[0]]
        data = values[1:]
    else:
        raise HTTPException(
            status_code=415,
            detail=f"Preview is not supported for '{suffix}'. Download it instead.",
        )

    from ..tools.excel_tools import _render_cell

    return {
        "name": path.name,
        "columns": headers,
        "total_rows": len(data),
        "rows": [[_render_cell(cell) for cell in row] for row in data[:rows]],
        "truncated": len(data) > rows,
        "download_url": f"/files/{path.name}",
    }


@app.get("/tools", summary="List the agent's toolbox")
async def tools() -> dict[str, Any]:
    settings = get_settings()
    registry = build_default_registry(Path(settings.tools_config))
    return {
        "enabled": [
            {
                "name": tool.name,
                "description": tool.effective_description,
                "tags": list(tool.tags),
                "input_schema": tool.input_schema(),
            }
            for tool in registry.enabled_tools()
        ],
        "disabled": [tool.name for tool in registry.all_tools() if not tool.enabled],
    }


@app.post("/runs", response_model=RunAccepted, status_code=202, summary="Start a run (non-blocking)")
async def start_run(request: RunRequest) -> RunAccepted:
    run_id = uuid.uuid4().hex[:12]
    handle = RunHandle(run_id, request)
    handle.loop = asyncio.get_running_loop()
    RUNS[run_id] = handle

    asyncio.get_running_loop().run_in_executor(None, _execute, handle)

    return RunAccepted(
        run_id=run_id,
        status="queued",
        events_url=f"/runs/{run_id}/events",
        result_url=f"/runs/{run_id}",
    )


@app.post("/runs/sync", summary="Start a run and block until it finishes")
async def start_run_sync(request: RunRequest) -> JSONResponse:
    run_id = uuid.uuid4().hex[:12]
    handle = RunHandle(run_id, request)
    handle.loop = asyncio.get_running_loop()
    RUNS[run_id] = handle

    await asyncio.get_running_loop().run_in_executor(None, _execute, handle)

    status_code = {"completed": 200, "partial": 207, "failed": 500}.get(handle.status, 500)
    return JSONResponse(
        status_code=status_code,
        content={
            "run_id": run_id,
            "status": handle.status,
            "error": handle.error,
            "result": handle.result,
        },
    )


@app.get("/runs", summary="List runs in this process")
async def list_runs() -> dict[str, Any]:
    return {
        "runs": [
            {"run_id": handle.run_id, "status": handle.status, "instruction": handle.request.instruction}
            for handle in RUNS.values()
        ]
    }


@app.get("/runs/{run_id}", summary="Fetch a run's status and full report")
async def get_run(run_id: str) -> dict[str, Any]:
    handle = RUNS.get(run_id)
    if handle is None:
        raise HTTPException(status_code=404, detail=f"Unknown run '{run_id}'")
    return {
        "run_id": handle.run_id,
        "status": handle.status,
        "instruction": handle.request.instruction,
        "error": handle.error,
        "result": handle.result,
        "event_count": len(handle.events),
    }


@app.get("/runs/{run_id}/events", summary="Live progress stream (SSE)")
async def stream_events(run_id: str) -> EventSourceResponse:
    handle = RUNS.get(run_id)
    if handle is None:
        raise HTTPException(status_code=404, detail=f"Unknown run '{run_id}'")

    queue = handle.subscribe()

    async def generator():
        try:
            while True:
                payload = await queue.get()
                if payload is None:
                    yield {"event": "done", "data": handle.status}
                    break
                yield {"event": payload["type"], "data": _json(payload)}
        finally:
            if queue in handle.queues:
                handle.queues.remove(queue)

    return EventSourceResponse(generator())


def _json(payload: dict[str, Any]) -> str:
    import json

    return json.dumps(payload, default=str, ensure_ascii=False)
