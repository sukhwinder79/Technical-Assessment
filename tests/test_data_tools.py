"""Data generation and format conversion."""

from __future__ import annotations

import csv
from pathlib import Path

import pytest

from agentic_sheets.errors import ToolError
from agentic_sheets.tools.base import ToolContext
from agentic_sheets.tools.data_tools import (
    BASE_COLUMNS,
    ConvertSpreadsheetTool,
    GenerateEmployeeCsvTool,
    ReadCsvPreviewTool,
    _coerce_number,
    generate_employee_rows,
)


# ---- pure generator --------------------------------------------------------


def test_generates_requested_row_count_and_required_columns():
    headers, rows = generate_employee_rows(25, seed=7)
    assert len(rows) == 25
    assert headers[: len(BASE_COLUMNS)] == BASE_COLUMNS


def test_meets_assessment_minimum_of_20_rows_by_default(ctx: ToolContext):
    result = GenerateEmployeeCsvTool().run(
        GenerateEmployeeCsvTool.args_model(), ctx  # all defaults
    )
    assert result["row_count"] >= 20


def test_same_seed_is_reproducible_and_different_seeds_differ():
    _, first = generate_employee_rows(15, seed=42)
    _, again = generate_employee_rows(15, seed=42)
    _, other = generate_employee_rows(15, seed=43)
    assert first == again
    assert first != other


def test_employee_ids_are_sequential_and_unique():
    _, rows = generate_employee_rows(30, seed=1)
    ids = [row[0] for row in rows]
    assert ids[0] == "EMP001"
    assert ids[-1] == "EMP030"
    assert len(set(ids)) == 30


def test_emails_are_unique_and_ascii_safe():
    _, rows = generate_employee_rows(200, seed=3)
    emails = [row[3] for row in rows]
    assert len(set(emails)) == len(emails)
    for email in emails:
        assert email.isascii() and email.count("@") == 1


def test_salaries_are_positive_integers():
    _, rows = generate_employee_rows(40, seed=5)
    for row in rows:
        assert isinstance(row[4], int) and row[4] > 0


def test_extended_columns_can_be_switched_off():
    headers, rows = generate_employee_rows(5, seed=1, include_extended=False)
    assert headers == BASE_COLUMNS
    assert all(len(row) == len(BASE_COLUMNS) for row in rows)


# ---- generate tool ---------------------------------------------------------


def test_csv_tool_writes_a_parsable_file_and_records_memory(ctx: ToolContext):
    tool = GenerateEmployeeCsvTool()
    result = tool.run(tool.args_model(row_count=22, filename="staff.csv", seed=11), ctx)

    path = Path(result["csv_path"])
    assert path.exists()
    assert result["row_count"] == 22

    with path.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    assert len(rows) == 23  # header + 22
    assert rows[0][: len(BASE_COLUMNS)] == BASE_COLUMNS

    # Working memory is what makes follow-up prompts work.
    assert ctx.memory.recall("last_csv_path") == str(path)
    assert ctx.memory.recall("last_csv_row_count") == 22


def test_csv_tool_forces_a_csv_extension(ctx: ToolContext):
    tool = GenerateEmployeeCsvTool()
    result = tool.run(tool.args_model(row_count=3, filename="oops.txt"), ctx)
    assert Path(result["csv_path"]).suffix == ".csv"


def test_relative_paths_are_confined_to_the_workspace(ctx: ToolContext):
    tool = GenerateEmployeeCsvTool()
    result = tool.run(tool.args_model(row_count=2, filename="nested/dir/out.csv"), ctx)
    assert Path(result["csv_path"]).is_relative_to(ctx.workspace)


def test_row_count_bounds_are_validated():
    with pytest.raises(Exception):
        GenerateEmployeeCsvTool.args_model(row_count=0)
    with pytest.raises(Exception):
        GenerateEmployeeCsvTool.args_model(row_count=99_999)


# ---- conversion ------------------------------------------------------------


@pytest.fixture
def sample_csv(ctx: ToolContext) -> Path:
    tool = GenerateEmployeeCsvTool()
    return Path(tool.run(tool.args_model(row_count=12, seed=2), ctx)["csv_path"])


def test_converts_csv_to_xlsx_with_typed_numbers(ctx: ToolContext, sample_csv: Path):
    from openpyxl import load_workbook

    tool = ConvertSpreadsheetTool()
    result = tool.run(tool.args_model(source_path=str(sample_csv), target_format="xlsx"), ctx)

    path = Path(result["output_path"])
    assert path.exists() and path.suffix == ".xlsx"

    workbook = load_workbook(path)
    sheet = workbook.active
    assert sheet.title == "Employees"
    assert sheet.max_row == 13  # header + 12
    assert sheet.freeze_panes == "A2"
    # Salary must be a number, not the string Excel would refuse to sum.
    salary_column = [c.value for c in sheet[1]].index("Salary") + 1
    assert isinstance(sheet.cell(row=2, column=salary_column).value, (int, float))
    workbook.close()


def test_converts_csv_to_ods(ctx: ToolContext, sample_csv: Path):
    tool = ConvertSpreadsheetTool()
    result = tool.run(tool.args_model(source_path=str(sample_csv), target_format="ods"), ctx)
    path = Path(result["output_path"])
    assert path.exists() and path.suffix == ".ods"
    assert path.stat().st_size > 0


def test_ods_round_trips_back_to_csv(ctx: ToolContext, sample_csv: Path):
    """ODS must be readable, not just writable — otherwise it is a dead end."""
    tool = ConvertSpreadsheetTool()
    ods = tool.run(tool.args_model(source_path=str(sample_csv), target_format="ods"), ctx)["output_path"]

    back = tool.run(
        tool.args_model(source_path=ods, target_format="csv", output_filename="from_ods.csv"), ctx
    )
    assert back["row_count"] == 12

    from agentic_sheets.tools.data_tools import read_csv as _read

    headers, rows = _read(Path(back["output_path"]))
    assert headers[:5] == BASE_COLUMNS
    assert len(rows) == 12


def test_reading_ods_preserves_values(ctx: ToolContext, sample_csv: Path):
    from agentic_sheets.tools.data_tools import read_csv as _read
    from agentic_sheets.tools.data_tools import read_ods

    original_headers, original_rows = _read(sample_csv)
    tool = ConvertSpreadsheetTool()
    ods = tool.run(tool.args_model(source_path=str(sample_csv), target_format="ods"), ctx)["output_path"]

    headers, rows = read_ods(Path(ods))
    assert headers == original_headers
    assert len(rows) == len(original_rows)
    assert rows[0][1] == original_rows[0][1]      # name survives
    assert str(rows[0][0]) == original_rows[0][0]  # employee id survives


def test_reading_a_corrupt_ods_is_a_clean_tool_error(ctx: ToolContext, tmp_path: Path):
    from agentic_sheets.tools.data_tools import read_ods

    broken = tmp_path / "broken.ods"
    broken.write_bytes(b"not an ods file at all")
    with pytest.raises(ToolError):
        read_ods(broken)


def test_round_trips_xlsx_back_to_csv(ctx: ToolContext, sample_csv: Path):
    tool = ConvertSpreadsheetTool()
    xlsx = tool.run(tool.args_model(source_path=str(sample_csv), target_format="xlsx"), ctx)["output_path"]
    back = tool.run(
        tool.args_model(source_path=xlsx, target_format="csv", output_filename="round_trip.csv"), ctx
    )
    assert back["row_count"] == 12


def test_conversion_rejects_a_missing_source(ctx: ToolContext):
    tool = ConvertSpreadsheetTool()
    with pytest.raises(ToolError) as exc:
        tool.run(tool.args_model(source_path="nope.csv", target_format="xlsx"), ctx)
    assert not exc.value.retryable  # a missing file is not worth retrying


def test_conversion_rejects_an_unsupported_format(ctx: ToolContext, tmp_path: Path):
    weird = tmp_path / "data.parquet"
    weird.write_bytes(b"\x00")
    tool = ConvertSpreadsheetTool()
    with pytest.raises(ToolError):
        tool.run(tool.args_model(source_path=str(weird), target_format="xlsx"), ctx)


# ---- preview ---------------------------------------------------------------


def test_preview_reports_columns_and_row_count(ctx: ToolContext, sample_csv: Path):
    tool = ReadCsvPreviewTool()
    result = tool.run(tool.args_model(csv_path=str(sample_csv), max_rows=3), ctx)
    assert result["row_count"] == 12
    assert result["columns"][:2] == ["Employee ID", "Name"]
    assert len(result["preview"]) == 4  # header + 3


# ---- helpers ---------------------------------------------------------------


@pytest.mark.parametrize(
    ("raw", "expected"),
    [("42", 42), ("3.5", 3.5), ("", ""), ("Sales", "Sales"), (None, None), (7, 7)],
)
def test_coerce_number(raw, expected):
    assert _coerce_number(raw) == expected
